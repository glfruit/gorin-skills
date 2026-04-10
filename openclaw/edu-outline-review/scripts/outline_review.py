from __future__ import annotations

import argparse
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


EXPECTED_HEADERS = ["项目名称", "任务名称", "模块/板块", "知识点/技能点", "主要内容概要"]
BANNED_VERBS = {
    "了解": "列举/陈述",
    "掌握": "说明/演示/完成",
    "熟悉": "辨识/说明",
    "培养": "形成/体现",
    "提高": "提升至可测行为",
    "增强": "体现/形成",
}
PLACEHOLDERS = ["TBD", "待定", "TODO", "XXX"]


@dataclass
class OutlineItem:
    row_no: int
    project: str
    task: str
    module: str
    unit: str
    point: str
    summary: str


@dataclass
class Heading:
    level: int
    title: str
    line_no: int


def normalize_text(value: str | None) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value)).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def compact_text(value: str | None) -> str:
    text = normalize_text(value).lower()
    text = re.sub(r"[\W_]+", "", text, flags=re.UNICODE)
    return text


def module_signature(value: str | None) -> str:
    text = normalize_text(value)
    text = re.sub(r"（.*?）|\(.*?\)", "", text)
    return compact_text(text)


def contains_match(needle: str, haystack: str) -> bool:
    needle_n = normalize_text(needle)
    haystack_n = normalize_text(haystack)
    if not needle_n:
        return False
    if needle_n in haystack_n:
        return True
    needle_c = compact_text(needle_n)
    haystack_c = compact_text(haystack_n)
    if not needle_c:
        return False
    if needle_c in haystack_c:
        return True
    if len(needle_c) >= 4:
        return similarity_ratio(needle_c, haystack_c) >= 0.72
    return False


def similarity_ratio(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    common = sum(1 for ch in set(a) if ch in set(b))
    return (2 * common) / (len(set(a)) + len(set(b)))


def detect_header_row(rows: list[tuple], template_type: str | None = None) -> int:
    preferred = 5 if template_type == "textbook" else 1
    candidates = []
    for idx, row in enumerate(rows[:15], start=1):
        normalized = [normalize_text(cell) for cell in row]
        score = sum(1 for header in EXPECTED_HEADERS if header in normalized)
        candidates.append((score, idx))
        if score >= 4:
            return idx
    if 1 <= preferred <= len(rows):
        return preferred
    candidates.sort(reverse=True)
    return candidates[0][1] if candidates else 1


def load_outline(excel_path: str, template_type: str | None = None) -> list[OutlineItem]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    preview_rows = list(ws.iter_rows(min_row=1, max_row=min(15, ws.max_row), values_only=True))
    header_row_no = detect_header_row(preview_rows, template_type=template_type)
    header_values = [normalize_text(v) for v in next(ws.iter_rows(min_row=header_row_no, max_row=header_row_no, values_only=True))]
    column_map = {name: idx for idx, name in enumerate(header_values) if name}

    def cell(row: tuple, name: str) -> str:
        idx = column_map.get(name)
        if idx is None or idx >= len(row):
            return ""
        return normalize_text(row[idx])

    items: list[OutlineItem] = []
    carry = {"项目名称": "", "任务名称": "", "模块/板块": "", "单元名称": ""}
    for row_no, row in enumerate(ws.iter_rows(min_row=header_row_no + 1, values_only=True), start=header_row_no + 1):
        if not any(value is not None and normalize_text(value) for value in row):
            continue
        for key in carry:
            value = cell(row, key)
            if value:
                carry[key] = value
        project = carry["项目名称"]
        task = carry["任务名称"]
        module = carry["模块/板块"]
        unit = carry["单元名称"]
        point = cell(row, "知识点/技能点")
        summary = cell(row, "主要内容概要")
        if point or summary:
            items.append(
                OutlineItem(
                    row_no=row_no,
                    project=project,
                    task=task,
                    module=module,
                    unit=unit,
                    point=point,
                    summary=summary,
                )
            )
        elif module or unit:
            items.append(
                OutlineItem(
                    row_no=row_no,
                    project=project,
                    task=task,
                    module=module,
                    unit=unit,
                    point="",
                    summary=summary,
                )
            )
    return items


def load_manuscript(manuscript_path: str) -> tuple[str, list[str], list[Heading]]:
    text = Path(manuscript_path).read_text(encoding="utf-8")
    lines = text.splitlines()
    headings: list[Heading] = []
    for idx, line in enumerate(lines, start=1):
        match = re.match(r"^(#{2,3})\s+(.*\S)\s*$", line)
        if match:
            headings.append(Heading(level=len(match.group(1)), title=normalize_text(match.group(2)), line_no=idx))
    return text, lines, headings


def build_sections(lines: list[str], headings: list[Heading]) -> dict[str, str]:
    sections: dict[str, str] = {}
    for i, heading in enumerate(headings):
        start = heading.line_no
        end = headings[i + 1].line_no - 1 if i + 1 < len(headings) else len(lines)
        body = "\n".join(lines[start:end]).strip()
        sections[heading.title] = body
    return sections


def find_heading_index(headings: list[Heading], target: str) -> int | None:
    target_n = normalize_text(target)
    for idx, heading in enumerate(headings):
        if contains_match(target_n, heading.title) or contains_match(heading.title, target_n):
            return idx
    return None


def extract_target_section(sections: dict[str, str], target_keywords: Iterable[str]) -> tuple[str, str]:
    for title, body in sections.items():
        for keyword in target_keywords:
            if contains_match(keyword, title):
                return title, body
    return "", ""


def map_outline_items(items: list[OutlineItem], manuscript_text: str, headings: list[Heading], sections: dict[str, str]):
    matches = []
    missing = []
    module_presence: dict[str, bool] = {}
    module_positions: dict[str, int] = {}

    for item in items:
        module_key = item.module or item.unit or "未命名板块"
        if module_key not in module_presence:
            module_index = find_heading_index(headings, item.module or item.unit)
            module_presence[module_key] = module_index is not None
            module_positions[module_key] = module_index if module_index is not None else 10**9

        search_targets = [item.point, item.summary, item.unit, item.module]
        search_targets = [target for target in search_targets if normalize_text(target)]
        matched = False
        evidence = ""
        matched_in_module = False

        module_title = item.module or item.unit
        module_text = ""
        if module_title:
            for title, body in sections.items():
                if contains_match(module_title, title) or contains_match(title, module_title):
                    module_text = f"{title}\n{body}"
                    break

        for target in search_targets:
            if contains_match(target, manuscript_text):
                matched = True
                evidence = target
            if module_text and contains_match(target, module_text):
                matched_in_module = True
                matched = True
                evidence = target
                break

        if not matched and item.point and module_text:
            point_tokens = [token for token in re.split(r"[、，,；;\s]+", item.point) if len(token) >= 2]
            if point_tokens and sum(1 for token in point_tokens if contains_match(token, module_text)) >= max(1, len(point_tokens) // 2):
                matched = True
                matched_in_module = True
                evidence = "、".join(point_tokens[:4])

        result = {
            "item": item,
            "matched": matched,
            "matched_in_module": matched_in_module,
            "module_present": module_presence[module_key],
            "evidence": evidence,
            "module_position": module_positions[module_key],
        }
        matches.append(result)
        if not matched:
            missing.append(result)

    return matches, missing, module_presence, module_positions


def analyze_structure(matches: list[dict]) -> list[str]:
    problems = []
    last_position = -1
    seen_modules = set()
    for record in matches:
        item = record["item"]
        module = item.module or item.unit or "未命名板块"
        module_key = module_signature(module)
        position = record["module_position"]
        if module_key in seen_modules or position == 10**9:
            continue
        if position < last_position:
            problems.append(f"- 板块顺序异常：`{module}` 在手稿中的位置早于大纲前序板块。")
        seen_modules.add(module_key)
        last_position = max(last_position, position)
    return problems


def analyze_goal_verbs(sections: dict[str, str]) -> list[dict[str, str]]:
    title, goal_text = extract_target_section(sections, ["任务目标"])
    if not goal_text:
        return []
    issues = []
    current_subsection = title
    for line in goal_text.splitlines():
        stripped = normalize_text(line)
        if not stripped:
            continue
        heading_match = re.match(r"^####\s+(.*)$", stripped)
        if heading_match:
            current_subsection = f"{title}-{heading_match.group(1)}"
            continue
        location = current_subsection
        for verb, suggestion in BANNED_VERBS.items():
            if verb in stripped:
                issues.append({"location": location, "verb": verb, "suggestion": suggestion, "content": stripped})
    return issues


def scan_placeholders(lines: list[str]) -> list[dict[str, str]]:
    hits = []
    for idx, line in enumerate(lines, start=1):
        for token in PLACEHOLDERS:
            if token in line:
                hits.append({"location": f"第{idx}行", "content": normalize_text(line)})
                break
    return hits


def render_table(rows: list[list[str]], headers: list[str]) -> str:
    lines = ["| " + " | ".join(headers) + " |", "|" + "|".join(["------"] * len(headers)) + "|"]
    if not rows:
        lines.append("| " + " | ".join(["无"] * len(headers)) + " |")
        return "\n".join(lines)
    for row in rows:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)


def determine_result(coverage_ratio: float, critical_count: int, important_count: int) -> str:
    if coverage_ratio < 0.75 or critical_count >= 3:
        return "❌ 退回"
    if coverage_ratio < 0.95 or critical_count > 0 or important_count > 0:
        return "⚠️ 需修复"
    return "✅ 通过"


def review_outline(manuscript_path: str, excel_path: str, template_type: str | None = None) -> str:
    outline_items = load_outline(excel_path, template_type=template_type)
    manuscript_text, lines, headings = load_manuscript(manuscript_path)
    sections = build_sections(lines, headings)
    matches, missing, module_presence, _ = map_outline_items(outline_items, manuscript_text, headings, sections)
    structure_problems = analyze_structure(matches)
    goal_issues = analyze_goal_verbs(sections)
    placeholders = scan_placeholders(lines)

    scored_matches = [record for record in matches if record["item"].point or record["item"].summary]
    scored_missing = [record for record in missing if record["item"].point or record["item"].summary]
    total_points = len(scored_matches)
    covered_points = sum(1 for record in scored_matches if record["matched"])
    missing_points = len(scored_missing)
    coverage_ratio = (covered_points / total_points) if total_points else 1.0

    module_pairs = []
    seen_module_keys = set()
    for item in outline_items:
        module = item.module or item.unit or "未命名板块"
        module_key = module_signature(module)
        if module_key in seen_module_keys:
            continue
        seen_module_keys.add(module_key)
        module_pairs.append((module, module_key))
    unique_modules = [module for module, _ in module_pairs]
    module_covered = sum(1 for module, module_key in module_pairs if module_presence.get(module, False) or module_presence.get(module_key, False))
    module_ratio = (module_covered / len(unique_modules)) if unique_modules else 1.0

    missing_rows = []
    for idx, record in enumerate(scored_missing, start=1):
        item = record["item"]
        missing_rows.append([
            str(idx),
            item.project or "-",
            item.task or "-",
            item.module or item.unit or "-",
            item.point or item.summary or "-",
            f"第{item.row_no}行",
        ])

    structure_text = "\n".join(structure_problems) if structure_problems else "- 未发现明显的板块顺序异常。"

    goal_rows = [
        [issue["location"], issue["verb"], issue["suggestion"]]
        for issue in goal_issues
    ]
    placeholder_rows = [
        [hit["location"], hit["content"]]
        for hit in placeholders
    ]

    important_count = len(goal_issues) + len(placeholders) + len(structure_problems)
    critical_count = missing_points
    result = determine_result(coverage_ratio, critical_count, important_count)

    uncovered_modules = [module for module, module_key in module_pairs if not (module_presence.get(module, False) or module_presence.get(module_key, False))]
    uncovered_module_text = "、".join(uncovered_modules) if uncovered_modules else "无"

    matched_examples = []
    for record in matches:
        if record["matched"] and record["item"].point:
            item = record["item"]
            matched_examples.append(
                f"- `{item.module or item.unit}` / `{item.point}` -> 命中依据：{record['evidence'] or '模糊匹配'}"
            )
        if len(matched_examples) >= 5:
            break
    if not matched_examples:
        matched_examples.append("- 未提取到可展示的命中样例。")

    review_time = datetime.now().strftime("%Y-%m-%d %H:%M")
    module_action = (
        f"- 模块覆盖率为 **{module_ratio:.1%}**，请优先补齐与大纲模块对应的二级或三级标题。"
        if module_ratio < 1.0
        else f"- 模块覆盖率为 **{module_ratio:.1%}**，模块层级已与大纲对齐。"
    )
    point_action = (
        f"- 缺失知识点共 **{missing_points}** 项，请逐条补写并回填到对应模块。"
        if missing_points > 0
        else "- 未发现缺失知识点，知识点覆盖达标。"
    )
    placeholder_action = (
        f"- 占位符命中 **{len(placeholders)}** 处，提交前必须清零。"
        if placeholders
        else "- 未发现占位符残留。"
    )

    report = f"""# 大纲符合度审查报告

## 基本信息
- **手稿**: {Path(manuscript_path).name}
- **大纲**: {Path(excel_path).name}
- **审查时间**: {review_time}

## 覆盖率统计
- **总知识点**: {total_points}
- **已覆盖**: {covered_points} ({coverage_ratio:.1%})
- **未覆盖**: {missing_points} ({(1 - coverage_ratio):.1%})
- **模块覆盖**: {module_covered}/{len(unique_modules)} ({module_ratio:.1%})
- **未命中模块**: {uncovered_module_text}

## 已覆盖样例
{chr(10).join(matched_examples)}

## 缺失知识点
{render_table(missing_rows, ['序号', '项目', '任务', '模块', '知识点', '大纲行号'])}

## 结构顺序问题
{structure_text}

## 教学目标动词问题
{render_table(goal_rows, ['位置', '禁用词', '建议替代'])}

## 占位符
{render_table(placeholder_rows, ['位置', '内容'])}

## 不一致项
{module_action}
{point_action}
- 任务目标禁用动词命中 **{len(goal_issues)}** 处，应改成可观察、可评价的行为动词。
{placeholder_action}

## 总评
- **结果**: {result}
- **覆盖率**: {coverage_ratio:.1%}
- **问题总数**: {critical_count + important_count}（Critical: {critical_count}, Important: {important_count}）
"""
    return report


def main() -> None:
    parser = argparse.ArgumentParser(description="Review manuscript markdown against Excel outline.")
    parser.add_argument("manuscript_path")
    parser.add_argument("excel_path")
    parser.add_argument("--template-type", default=None)
    parser.add_argument("--output", default=None)
    args = parser.parse_args()

    report = review_outline(args.manuscript_path, args.excel_path, template_type=args.template_type)
    print(report)
    if args.output:
        Path(args.output).write_text(report, encoding="utf-8")


if __name__ == "__main__":
    main()
